#!/usr/bin/env python3
"""
Build the live-music venue scouting spreadsheet from the 3 web-research
agent outputs (Aug 2026). Output: /opt/gigsfill/gigsfill_scouting_list.xlsx.

Columns: Venue | City | Address | Phone | Website | Entertainment | Evidence

Freeze top row, bold header, alternating row background, autosized-ish
column widths. Ready to hand to the sales / outreach effort.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Every row is a tuple:
# (venue, city, address, phone, website, entertainment, evidence)
VENUES = [
    # ── Agoura Hills ─────────────────────────────────────────────────
    ("The Canyon Club & Special Events Center", "Agoura Hills", "28912 Roadside Dr, Agoura Hills, CA 91301",
     "(818) 879-5016", "canyonclub.net",
     "Live music, comedy",
     "Dedicated 1,200-cap dinner-theater music venue in Whizin Center; nightly touring/tribute acts."),
    ("Tavern Tomoko & Ladyface Brewery", "Agoura Hills", "29281 Agoura Rd, Agoura Hills, CA 91301",
     "(818) 477-4566", "ladyfaceale.com",
     "Trivia (comedian-hosted)",
     "Live multi-media trivia every Wednesday at 6:30pm, hosted by local comedians."),
    ("The Old Place", "Agoura Hills", "29983 Mulholland Hwy, Agoura Hills, CA 91301",
     "", "oldplacecornell.com",
     "Live acoustic music",
     "Acoustic musicians stroll through the restaurant — best on Thursday or Friday nights."),
    ("Adobe Cantina", "Agoura Hills", "29100 Agoura Rd, Agoura Hills, CA 91301",
     "(818) 991-3474", "adobecantina.com",
     "Live music, mariachi",
     "Listed among top Agoura Hills Mexican restaurants with live mariachi."),
    ("Blue Table Wine & Cheese Bar", "Agoura Hills", "28912 Roadside Dr #101, Agoura Hills, CA 91301",
     "(818) 597-2583", "bluetable.net",
     "Live music",
     "Public events calendar of live music dates; on GigTown as a live-music venue."),

    # ── Calabasas ────────────────────────────────────────────────────
    ("Sagebrush Cantina", "Calabasas", "23527 Calabasas Rd, Calabasas, CA 91302",
     "(818) 222-6062", "sagebrushcantina.com",
     "Live music, DJs, karaoke",
     "Thursday karaoke, Fri/Sat bands + dancing, Sunday afternoon music."),
    ("Malibu Canyon Bar & Grill (Cambria Hotel)", "Calabasas", "26400 Rondell St, Calabasas, CA 91302",
     "(818) 878-9950", "cambriacalabasas.com",
     "Live music, DJ",
     "Weekly live music every Taco Tuesday and Friday 7-9pm plus DJ dates."),
    ("The Commons at Calabasas (outdoor series)", "Calabasas", "4799 Commons Way, Calabasas, CA 91302",
     "(818) 637-8922", "shopcommons.com",
     "Live music (outdoor series)",
     "Live performers every Thursday 5:30-8:30pm plus Saturday music; hosted on Toscanova/Marmalade/King's Fish patios."),

    # ── Westlake Village ─────────────────────────────────────────────
    ("NABU Wines", "Westlake Village", "2649 Townsgate Rd, Suite 200, Westlake Village, CA 91361",
     "(805) 778-1100", "nabuwines.com",
     "Live music, trivia",
     "King Trivia Thu 7-10pm, live bands Fri/Sat 8-11pm, Sunday Tunes acoustic 3-6pm."),
    ("The Stonehaus", "Westlake Village", "32039 Agoura Rd, Westlake Village, CA 91361",
     "(818) 483-1152", "the-stonehaus.com",
     "Live music",
     "Recurring Spanish Nights with live music Fri 7-10pm, Sat 7-10pm, Sun 3-7pm."),
    ("Bogie's at the Westlake Village Inn", "Westlake Village", "32001 Agoura Rd, Westlake Village, CA 91361",
     "(818) 889-2394", "bogies-bar.com",
     "Live music (concert series)",
     "Dedicated Live Music Series — Colbie Caillat, Tower of Power etc. on 2026-27 calendar."),
    ("Louie's (Westlake Village Inn)", "Westlake Village", "32001 Agoura Rd, Westlake Village, CA 91361",
     "(818) 889-2394", "clublouies.com",
     "Live music, DJ / club nights",
     "Live music and club nights Wed-Sat; publishes calendar of bands + DJs."),
    ("Naughty Pine Brewing Co.", "Westlake Village", "766 Lakefield Dr, Suite A, Westlake Village, CA 91361",
     "(805) 906-2140", "naughtypinebrewingco.com",
     "Trivia, live music",
     "Trivia Thursdays 7-9pm plus live music most evenings per Conejo Valley Guide."),
    ("14 Cannons Brewery", "Westlake Village", "31125 Via Colinas #907, Westlake Village, CA 91362",
     "(818) 264-4213", "14cannons.com",
     "Live music, comedy, R&B brunch",
     "Aug 2026 events: live music (Aug 21/28/29), comedy (Aug 20/29), R&B sing-along brunch (Aug 22)."),
    ("Prosperous Penny (Four Seasons Westlake Village)", "Westlake Village", "2 Dole Dr, Westlake Village, CA 91362",
     "(818) 575-3000", "fourseasons.com/westlakevillage",
     "Live jazz",
     "Live jazz Tuesdays & Thursdays 8-10pm per Four Seasons dining page."),

    # ── Thousand Oaks (incl. Newbury Park) ───────────────────────────
    ("Crown & Anchor", "Thousand Oaks", "2891 E Thousand Oaks Blvd, Thousand Oaks, CA 91362",
     "(805) 497-0070", "crownnanchor.com",
     "Live music, trivia",
     "British pub — live music on the patio Fri/Sat + Sun afternoons and weekly trivia."),
    ("Pedals & Pints Brewing Co.", "Thousand Oaks", "156 W Hillcrest Dr, Thousand Oaks, CA 91360",
     "(805) 551-0547", "pedalsandpintsbrewing.com",
     "Live music, trivia",
     "Live music every Saturday and Trivia on Tap every Wednesday at 7pm."),
    ("Boney Mountain Pizza Co.", "Newbury Park", "722 N Wendy Dr, Newbury Park, CA 91320",
     "(805) 498-7200", "boneymountainpizzaco.com",
     "Trivia",
     "Trivia Night every Tuesday at 7pm."),
    ("Bottle and Pint", "Newbury Park", "1714 Newbury Rd, Newbury Park, CA 91320",
     "(805) 480-9500", "",
     "Trivia, live music",
     "King Trivia venue; Yelp reviews note weekly trivia and live music on Fridays."),
    ("Breakers Sports Bar & Grill", "Thousand Oaks", "398 N Moorpark Rd, Thousand Oaks, CA 91360",
     "(805) 494-8454", "breakersto.com",
     "Live music, karaoke",
     "Live bands every weekend, karaoke nights."),
    ("Azar's Sports Bar & Grill", "Newbury Park", "2215 Michael Dr, Newbury Park, CA 91320",
     "(805) 498-2365", "azarsportsbar.com",
     "Karaoke, live music, DJs",
     "Karaoke Tue/Fri, live bands + DJs on weekends."),
    ("PJ's Sports Pub", "Thousand Oaks", "417 E Avenida De Los Arboles, Thousand Oaks, CA 91360",
     "", "",
     "Karaoke, live music",
     "Karaoke Sun/Mon nights, live bands on select weekends."),

    # ── Camarillo ────────────────────────────────────────────────────
    ("The Manhattan of Camarillo", "Camarillo", "5800 Santa Rosa Rd, Ste 140, Camarillo, CA 93012",
     "(805) 388-5550", "themanhattanofcamarillo.com",
     "Live music, trivia",
     "Live music Thu-Sun, Trivia on Thursdays per venue site."),
    ("Institution Ale Co.", "Camarillo", "3841 Mission Oaks Blvd, Camarillo, CA 93012",
     "(805) 482-3777", "institutionales.com",
     "Trivia, live music",
     "King Trivia Tuesdays 7pm, Live Music Thursdays in beer garden."),
    ("Topa Topa Brewing Co. (Camarillo taproom)", "Camarillo", "2024 Ventura Blvd, Camarillo, CA 93010",
     "(805) 702-4091", "topatopa.beer",
     "Live music",
     "Rotating live music at the Camarillo taproom per Happenings calendar."),
    ("House of Bamboo", "Camarillo", "2227 Ventura Blvd, Camarillo, CA 93010",
     "", "houseofbamboo.com",
     "Live music, DJs",
     "Tiki lounge — cocktails, food, and music; IG shows DJ/live sets. Reservations via OpenTable."),
    ("The Amendment", "Camarillo", "350 N Lantana St, Ste 7, Camarillo, CA 93010",
     "(805) 201-6300", "facebook.com/theamendmentbar",
     "Live music, DJs",
     "Speakeasy hosts live music sessions and DJs."),
    ("Twenty 88 Food and Drink", "Camarillo", "2088 Ventura Blvd, Camarillo, CA 93010",
     "(805) 388-2088", "twenty88.com",
     "Live music",
     "Live music on Fridays and select nights per GigTown/Yelp."),

    # ── Moorpark ─────────────────────────────────────────────────────
    ("Enegren Brewing Company", "Moorpark", "444 Zachary St #120, Moorpark, CA 93021",
     "(805) 552-0602", "enegrenbrewing.com",
     "Live music, trivia",
     "Biergarten hosts live music, trivia nights, and community events on a regular cadence."),
    ("Lucky Fools Pub", "Moorpark", "75 E High St, Moorpark, CA 93021",
     "(805) 532-1500", "luckyfoolspub.com",
     "Live music, karaoke",
     "Dedicated Live Music page on venue site; KaraMap listing confirms karaoke."),
    ("The Ranch Restaurant & Bar", "Moorpark", "15187 Tierra Rejada Rd, Moorpark, CA 93021",
     "(805) 531-9300", "",
     "Live music",
     "Live music, best nights on Friday per Yelp reviews (on Tierra Rejada golf course)."),
    ("Freda's Fine Food & Drink", "Moorpark", "233 E High St, Moorpark, CA 93021",
     "(805) 217-4548", "fredasmoorpark.com",
     "Live music",
     "Live jazz on the patio incl. Moorpark HS jazz band per Yelp + Moorpark Acorn."),

    # ── Simi Valley ──────────────────────────────────────────────────
    ("The Arena Grill & Lounge", "Simi Valley", "999 Enchanted Way, Simi Valley, CA 93065",
     "(805) 915-1528", "arenasimi.com",
     "Live music, DJ",
     "Live music Mon/Tue/Wed 8pm plus Music By Cheps DJ nights per venue site."),
    ("Cork & Batter Roadhouse", "Simi Valley", "1747 Simi Town Center Way, Simi Valley, CA 93065",
     "(805) 210-2290", "corkandbatter.com",
     "Live music, karaoke",
     "Weekend live music (rock/country/blues) Fri/Sat, line dancing, karaoke."),
    ("Nectar of the Dogs Wine", "Simi Valley", "791 Chambers Ln Ste 130, Simi Valley, CA 93065",
     "(805) 624-7101", "nectarofthedogswine.com",
     "Live music (acoustic), trivia",
     "Trivia Thursdays 7pm; live acoustic weekends per venue events page."),
    ("Rock N Roll Pizza Bar (Harley's Valley Bowl)", "Simi Valley", "5255 Cochran St, Simi Valley, CA 93063",
     "(805) 584-2695", "harleysbowl.com/rocknrollpizzabar",
     "Live music, comedy, karaoke",
     "Weekly live bands, comedy Wed, karaoke Tue, reggae Sun."),
    ("The Hive Bar and Lounge", "Simi Valley", "2780 Tapo Canyon Rd Ste B4, Simi Valley, CA 93063",
     "(805) 210-2267", "",
     "DJs, live music, karaoke",
     "DJs weekends, karaoke nights, occasional live music per Visit Simi Valley + Yelp."),
    ("JR's Comedy Club (inside Junkyard Cafe)", "Simi Valley", "2585 Cochran St, Simi Valley, CA 93065",
     "(805) 426-6117", "tocomedy.com",
     "Comedy",
     "Standalone comedy club — regular weekend stand-up bookings."),
    ("Chuy's Baja Grill (Simi West)", "Simi Valley", "1397 E Los Angeles Ave, Simi Valley, CA 93065",
     "(805) 582-4897", "chuysbajagrill.com",
     "Trivia",
     "Listed as a King Trivia weekly host venue."),

    # ── Oxnard ───────────────────────────────────────────────────────
    ("Copper Blues Rock Pub & Kitchen", "Oxnard", "591 Collection Blvd, Oxnard, CA 93036",
     "(805) 457-5551", "copperblueslive.com/locations/oxnard",
     "Live music, DJs",
     "Live rock/tribute bands, DJs Fri/Sat, salsa & mariachi Sun."),
    ("Waterside Restaurant & Wine Bar", "Oxnard", "3500 S Harbor Blvd Ste 111, Oxnard, CA 93035",
     "(805) 985-4677", "watersidechannelislands.com",
     "Live music",
     "Live entertainment Thu/Fri/Sat evenings + Sun brunch."),
    ("The Lookout Bar & Grill", "Oxnard", "2800 Harbor Blvd, Oxnard, CA 93035",
     "(805) 985-9300", "",
     "Live music, trivia, karaoke, open mic",
     "All four listed on Visit Oxnard directory + badslava trivia directory."),
    ("The Thirsty Ox", "Oxnard", "640 N Ventura Rd, Oxnard, CA 93030",
     "(805) 307-1015", "",
     "Trivia, live music, DJs",
     "Trivia Wed, DJ Fri/Sat, live music per King Trivia + Yelp."),
    ("Wagon Wheel Brewing Company", "Oxnard", "2601 Wagon Wheel Rd Ste 40, Oxnard, CA 93036",
     "", "wagonwheelbrewingcompany.com",
     "Live music, trivia",
     "Live music, trivia, private events, community gatherings per own site."),
    ("1901 Speakeasy", "Oxnard", "740 S B St, Oxnard, CA 93030",
     "(805) 486-6878", "the1901.com",
     "Live music (jazz/swing)",
     "Multi-night live music per Visit Oxnard. NOTE: Visit Oxnard flagged 'Temporarily Closed' — confirm status before outreach."),

    # ── Ventura ──────────────────────────────────────────────────────
    ("Bright Spark Brewing & Restaurant", "Ventura", "4561 Market St, Ventura, CA 93003",
     "(805) 322-8884", "brightsparkbrewing.com",
     "Trivia, music bingo",
     "Trivia Wed 6:30pm, Music Bingo Thu 7pm — both weekly."),
    ("Topa Topa Brewing (Downtown)", "Ventura", "104 E Thompson Blvd, Ventura, CA 93001",
     "(805) 628-9255", "topatopa.beer",
     "Trivia",
     "Independent trivia Thursday 7pm per Ventura trivia guide."),
    ("Ventura Music Hall", "Ventura", "1888 E Thompson Blvd, Ventura, CA 93001",
     "(805) 667-8802", "venturamusichall.com",
     "Live music",
     "Dedicated 635-cap live venue; touring + local acts, full bar and kitchen."),
    ("Bombay Bar & Grill", "Ventura", "143 S California St, Ventura, CA 93001",
     "(805) 643-4404", "bombayventura.com",
     "Live music, DJs",
     "Live bands & DJs Wed-Sun, two stages, two dance floors."),
    ("The Cave (Ventura Wine Co.)", "Ventura", "4435 McGrath St Ste 301, Ventura, CA 93003",
     "(805) 642-9449", "thecaveventura.com",
     "Live music",
     "Live music select nights per venue site."),
    ("Cafe Fiore", "Ventura", "66 S California St, Ventura, CA 93001",
     "(805) 653-1266", "cafefioreventura.com",
     "Live music (jazz + Latin/reggae/top 40)",
     "Martini Lounge features jazz Tue nights + rotating bands."),
    ("Leashless Brewing", "Ventura", "585 E Thompson Blvd, Ventura, CA 93001",
     "(805) 628-9474", "leashlessbrewing.com",
     "Live music",
     "Music & Events page — regular live music + food trucks; on GigTown live-music venue index."),
]

HEADERS = ["Venue", "City", "Address", "Phone", "Website", "Entertainment", "Evidence"]
COL_WIDTHS = [34, 18, 44, 16, 32, 32, 66]

OUT = "/opt/gigsfill/gigsfill_scouting_list.xlsx"


def _apply_header(ws):
    fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left")


def _apply_row(ws, row_idx, values):
    alt = row_idx % 2 == 0  # zebra stripe
    fill = PatternFill(start_color="F5F5FA" if alt else "FFFFFF", end_color="F5F5FA" if alt else "FFFFFF",
                       fill_type="solid")
    thin = Side(style="thin", color="E1E1EA")
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        cell.fill = fill
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Venues"

    _apply_header(ws)
    ws.freeze_panes = "A2"

    # Sort by city then venue so each city clusters together.
    rows = sorted(VENUES, key=lambda r: (r[1], r[0]))
    for i, row in enumerate(rows, start=2):
        _apply_row(ws, i, row)

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 24
    # Bump every content row so wrapped evidence text has breathing room
    for i in range(2, len(rows) + 2):
        ws.row_dimensions[i].height = 42

    wb.save(OUT)
    print(f"Wrote {len(rows)} venues to {OUT}")


if __name__ == "__main__":
    main()
